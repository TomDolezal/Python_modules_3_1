"""
test_cell_model_io.py
=======================

Simple, dependency-light sanity tests for cell_model_io.py.
Runnable directly:

    python test_cell_model_io.py

Uses the sample workbook shipped alongside this repo
(sample_data/sample_cell_model.xlsx).
"""

import sys
import os
import shutil
import tempfile

import numpy as np
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from cell_model_io import load_cell_model_xlsx, make_sim_cell_model
from profile_generation import run_battery_model
from _pg_helpers import apply_default_options, apply_default_datasheet

SAMPLE_XLSX = os.path.join(os.path.dirname(__file__), "..", "sample_data",
                            "sample_cell_model.xlsx")


def _require_sample():
    if not os.path.exists(SAMPLE_XLSX):
        raise FileNotFoundError(
            f"Sample cell model workbook not found at {SAMPLE_XLSX}. "
            "Place a copy there (or edit SAMPLE_XLSX at the top of this file) to run these tests."
        )


def test_load_matches_raw_values():
    _require_sample()
    cm = load_cell_model_xlsx(SAMPLE_XLSX)

    assert np.isclose(cm.soc[0], 0.0) and np.isclose(cm.soc[-1], 1.0)
    assert cm.ocv.shape == cm.soc.shape
    assert cm.r0.shape == cm.soc.shape
    assert np.isclose(cm.ocv[0], 2.007235714286429)
    assert np.isclose(cm.capacity_Ah, 3.35)
    assert cm.datasheet_meta.get("type") == "prismatic"
    print("PASS: loaded values match the raw workbook exactly")


def test_end_to_end_simulation_with_real_cell_model():
    _require_sample()
    cm = load_cell_model_xlsx(SAMPLE_XLSX)
    sim_model = make_sim_cell_model(cm)

    rng = np.random.default_rng(3)
    n = 300
    t = np.arange(n)
    cell_power = 1.5 * np.sin(2 * np.pi * t / 60) + rng.normal(0, 0.3, n)

    opts = apply_default_options({"scale_divisor": 1.0, "dt": 1.0, "T_cell_degC": 25,
                                   "charging_positive": False})
    datasheet = apply_default_datasheet({"max_current_charge_A": 5, "max_current_discharge_A": 5})

    res = run_battery_model({"power": cell_power, "soc_init": 60.0}, "power_only",
                             sim_model, datasheet, opts)
    assert res.is_valid, res.message
    assert 0 < res.soc.min() and res.soc.max() < 100
    print("PASS: real cell model runs end-to-end through run_battery_model")


def test_missing_required_sheet():
    _require_sample()
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "missing_sheet.xlsx")
        shutil.copy(SAMPLE_XLSX, path)
        wb = openpyxl.load_workbook(path)
        del wb["R0"]
        wb.save(path)
        try:
            load_cell_model_xlsx(path)
            raise AssertionError("should have raised")
        except ValueError as e:
            assert "R0" in str(e)
    print("PASS: missing required sheet raises a clear ValueError")


def test_extra_r0_column_rejected():
    _require_sample()
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "extra_col.xlsx")
        shutil.copy(SAMPLE_XLSX, path)
        wb = openpyxl.load_workbook(path)
        wb["R0"].cell(row=1, column=3, value="R0_extra")
        wb["R0"].cell(row=2, column=3, value=0.1)
        wb.save(path)
        try:
            load_cell_model_xlsx(path)
            raise AssertionError("should have raised")
        except ValueError as e:
            assert "resistance column" in str(e)
    print("PASS: more than one R0 column raises a clear ValueError")


def test_mismatched_soc_grid():
    _require_sample()
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "bad_soc.xlsx")
        shutil.copy(SAMPLE_XLSX, path)
        wb = openpyxl.load_workbook(path)
        wb["R0"].cell(row=2, column=1, value=0.001)  # tweak SOC away from OCV sheet's value
        wb.save(path)
        try:
            load_cell_model_xlsx(path)
            raise AssertionError("should have raised")
        except ValueError as e:
            assert "SOC" in str(e)
    print("PASS: mismatched SOC grid between OCV and R0 raises a clear ValueError")


def test_missing_capacity_row():
    _require_sample()
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "no_capacity.xlsx")
        shutil.copy(SAMPLE_XLSX, path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Meta"]
        for row in ws.iter_rows():
            if row[0].value == "capacity_Ah":
                ws.delete_rows(row[0].row)
                break
        wb.save(path)
        try:
            load_cell_model_xlsx(path)
            raise AssertionError("should have raised")
        except ValueError as e:
            assert "capacity_Ah" in str(e)
    print("PASS: missing capacity_Ah row raises a clear ValueError")


def test_comma_decimal_tolerance():
    """A hand-edited cell stored as text with a comma decimal (e.g. '3,6')
    should still parse correctly."""
    _require_sample()
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "comma_decimal.xlsx")
        shutil.copy(SAMPLE_XLSX, path)
        wb = openpyxl.load_workbook(path)
        wb["Meta"].cell(row=2, column=2, value="3,6")  # voltage_V as comma-decimal text
        wb.save(path)
        result = load_cell_model_xlsx(path)
        assert np.isclose(result.datasheet_meta["voltage_V"], 3.6)
    print("PASS: comma-decimal text values are parsed correctly")


def test_datasheet_current_limits_from_meta():
    """max_current_charge_A / max_current_discharge_A, if present on the
    Meta sheet, should be surfaced as a ready-to-use `datasheet` dict --
    absent fields default to unconstrained (inf), matching
    profile_generation's own defaulting."""
    _require_sample()

    # The base sample already defines both limits on the Meta sheet.
    result = load_cell_model_xlsx(SAMPLE_XLSX)
    assert np.isclose(result.datasheet["max_current_charge_A"], 21.0), \
        f"expected 21.0, got {result.datasheet['max_current_charge_A']!r} (datasheet={result.datasheet})"
    assert np.isclose(result.datasheet["max_current_discharge_A"], 21.0), \
        f"expected 21.0, got {result.datasheet['max_current_discharge_A']!r} (datasheet={result.datasheet})"
    assert "max_current_charge_A" not in result.datasheet_meta, \
        f"max_current_charge_A leaked into datasheet_meta: {result.datasheet_meta}"

    with tempfile.TemporaryDirectory() as tmp:
        # Neither present -> both default to inf
        path0 = os.path.join(tmp, "no_limits.xlsx")
        shutil.copy(SAMPLE_XLSX, path0)
        wb0 = openpyxl.load_workbook(path0)
        rows_to_delete = [row[0].row for row in wb0["Meta"].iter_rows()
                          if row[0].value in ("max_current_charge_A", "max_current_discharge_A")]
        for row_idx in sorted(rows_to_delete, reverse=True):
            wb0["Meta"].delete_rows(row_idx)
        wb0.save(path0)
        r0 = load_cell_model_xlsx(path0)
        assert r0.datasheet["max_current_charge_A"] == np.inf, \
            f"expected inf, got {r0.datasheet['max_current_charge_A']!r} (datasheet={r0.datasheet})"
        assert r0.datasheet["max_current_discharge_A"] == np.inf, \
            f"expected inf, got {r0.datasheet['max_current_discharge_A']!r} (datasheet={r0.datasheet})"

        # Both present
        path = os.path.join(tmp, "with_limits.xlsx")
        shutil.copy(path0, path)
        wb = openpyxl.load_workbook(path)
        wb["Meta"].append(["max_current_charge_A", 10.5])
        wb["Meta"].append(["max_current_discharge_A", 21.0])
        wb.save(path)
        r2 = load_cell_model_xlsx(path)
        assert np.isclose(r2.datasheet["max_current_charge_A"], 10.5), \
            f"expected 10.5, got {r2.datasheet['max_current_charge_A']!r} (full datasheet={r2.datasheet})"
        assert np.isclose(r2.datasheet["max_current_discharge_A"], 21.0), \
            f"expected 21.0, got {r2.datasheet['max_current_discharge_A']!r} (full datasheet={r2.datasheet})"
        assert "max_current_charge_A" not in r2.datasheet_meta, \
            f"max_current_charge_A leaked into datasheet_meta: {r2.datasheet_meta}"
        assert r2.datasheet_meta.get("type") == "prismatic", \
            f"expected type='prismatic', got {r2.datasheet_meta.get('type')!r} (full datasheet_meta={r2.datasheet_meta})"

        # Only one present -> the other still defaults to inf
        path2 = os.path.join(tmp, "partial_limits.xlsx")
        shutil.copy(path0, path2)
        wb2 = openpyxl.load_workbook(path2)
        wb2["Meta"].append(["max_current_charge_A", 8.0])
        wb2.save(path2)
        r3 = load_cell_model_xlsx(path2)
        assert np.isclose(r3.datasheet["max_current_charge_A"], 8.0), \
            f"expected 8.0, got {r3.datasheet['max_current_charge_A']!r} (full datasheet={r3.datasheet})"
        assert r3.datasheet["max_current_discharge_A"] == np.inf, \
            f"expected inf, got {r3.datasheet['max_current_discharge_A']!r} (full datasheet={r3.datasheet})"

    print("PASS: max_current_charge_A / max_current_discharge_A round-trip via the Meta sheet")


def test_datasheet_from_excel_enforced_end_to_end():
    """The datasheet dict loaded from Excel should actually be enforced
    by run_battery_model when use_current_limit=True."""
    _require_sample()
    with tempfile.TemporaryDirectory() as tmp:
        path = os.path.join(tmp, "tight_limits.xlsx")
        shutil.copy(SAMPLE_XLSX, path)
        wb = openpyxl.load_workbook(path)
        wb["Meta"].append(["max_current_charge_A", 0.3])
        wb["Meta"].append(["max_current_discharge_A", 0.3])
        wb.save(path)

        result = load_cell_model_xlsx(path)
        sim_model = make_sim_cell_model(result)

        rng = np.random.default_rng(3)
        n = 200
        cell_power = 1.5 * np.sin(2 * np.pi * np.arange(n) / 60) + rng.normal(0, 0.3, n)
        opts = apply_default_options({"scale_divisor": 1.0, "dt": 1.0, "T_cell_degC": 25,
                                       "charging_positive": False, "use_current_limit": True})

        res = run_battery_model({"power": cell_power, "soc_init": 60.0}, "power_only",
                                 sim_model, result.datasheet, opts)
        assert not res.is_valid
        assert "limit exceeded" in res.message
    print("PASS: Excel-loaded datasheet current limits are enforced end-to-end")


def run_all():
    import traceback

    tests = [
        test_load_matches_raw_values,
        test_end_to_end_simulation_with_real_cell_model,
        test_missing_required_sheet,
        test_extra_r0_column_rejected,
        test_mismatched_soc_grid,
        test_missing_capacity_row,
        test_comma_decimal_tolerance,
        test_datasheet_current_limits_from_meta,
        test_datasheet_from_excel_enforced_end_to_end,
    ]

    failures = []
    for t in tests:
        try:
            t()
        except Exception as e:
            failures.append((t.__name__, e))
            print(f"FAIL: {t.__name__}")
            traceback.print_exc()
            print()

    print()
    if failures:
        print(f"{len(failures)}/{len(tests)} tests FAILED")
        sys.exit(1)
    else:
        print(f"All {len(tests)} tests PASSED")


if __name__ == "__main__":
    run_all()
