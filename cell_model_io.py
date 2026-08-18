"""
cell_model_io
==============

Loader for the SOC-OCV-R0 cell-model workbook schema: just
`Meta` / `OCV` / `R0`, with no temperature axis --

    OCV sheet:  SOC, OCV_V
    R0 sheet:   SOC, R0_<label>          (exactly one resistance column)
    Meta sheet: label/value pairs, capacity_Ah required

Usage
-----
    from cell_model_io import load_cell_model_xlsx, make_sim_cell_model
    from profile_generation import run_battery_model, tune_battery_model

    cm = load_cell_model_xlsx("cell_model.xlsx")
    sim_model = make_sim_cell_model(cm)  # eta not in the workbook -> falls
                                          # back to opts['eta_coulomb_ef']

    # sim_model has the same keys/call signatures run_battery_model /
    # tune_battery_model expect, so those functions work directly with it.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Optional

import numpy as np
import openpyxl

from _pg_helpers import apply_default_datasheet, make_clamped_interp1d

REQUIRED_SHEETS = ("Meta", "OCV", "R0")
DATASHEET_FIELDS = ("max_current_charge_A", "max_current_discharge_A")


@dataclass
class CellModelFromExcel:
    """Everything loaded from a SOC-OCV-R0-only workbook."""
    soc: np.ndarray                # normalized 0..1
    ocv: np.ndarray                # V, aligned with soc
    r0: np.ndarray                 # ohm, aligned with soc
    r0_column_label: str           # e.g. "R0_25C", for reference/logging only
    capacity_Ah: float
    datasheet: dict = field(default_factory=dict)
    datasheet_meta: dict = field(default_factory=dict)
    source_path: str = ""


def _to_float(value) -> float:
    if value is None:
        raise ValueError("Empty cell where a numeric value was expected.")
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip().replace(",", ".")
        try:
            return float(s)
        except ValueError:
            raise ValueError(f"Could not parse '{value}' as a number.")
    raise ValueError(f"Unexpected cell type for numeric value: {type(value)}")


def _read_meta_sheet(ws) -> dict:
    """Reads label/value rows."""
    fields = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] is None:
            continue
        label = str(row[0]).strip()
        if label == "Var1":
            continue  # header row
        if len(row) > 1 and row[1] is not None:
            val = row[1]
            try:
                val = _to_float(val)
            except ValueError:
                val = str(val).strip()
            fields[label] = val
    return fields


def _read_ocv_sheet(ws) -> tuple[np.ndarray, np.ndarray]:
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h is not None else "" for h in rows[0]]

    required = {"SOC", "OCV_V"}
    if not required.issubset(set(header)):
        raise ValueError(
            f"OCV sheet must contain columns {sorted(required)}, found {header}."
        )
    idx = {name: header.index(name) for name in required}
    data_rows = [r for r in rows[1:] if r[idx["SOC"]] is not None]

    soc = np.array([_to_float(r[idx["SOC"]]) for r in data_rows])
    ocv = np.array([_to_float(r[idx["OCV_V"]]) for r in data_rows])

    if not (np.all(soc >= 0) and np.all(soc <= 1)):
        raise ValueError("OCV sheet: SOC values must be in the interval [0, 1].")
    if not np.all(np.diff(soc) > 0):
        raise ValueError("OCV sheet: SOC values must be strictly increasing.")

    return soc, ocv


def _read_r0_sheet(ws, expected_soc: np.ndarray) -> tuple[np.ndarray, str]:
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() if h is not None else "" for h in rows[0]]

    if not header or header[0] != "SOC":
        raise ValueError(f"R0 sheet: first column must be 'SOC', found header {header}.")

    r0_cols = [h for h in header[1:] if h]
    if len(r0_cols) != 1:
        raise ValueError(
            f"R0 sheet must have exactly one resistance column, "
            f"found {len(r0_cols)}: {r0_cols}."
        )
    r0_label = r0_cols[0]
    r0_col_idx = header.index(r0_label)

    data_rows = [r for r in rows[1:] if r[0] is not None]
    if len(data_rows) != len(expected_soc):
        raise ValueError(
            f"R0 sheet has {len(data_rows)} SOC rows, but the OCV sheet defines "
            f"{len(expected_soc)} SOC points. These must match."
        )

    soc_r0 = np.array([_to_float(r[0]) for r in data_rows])
    if not np.allclose(soc_r0, expected_soc):
        raise ValueError(
            "R0 sheet: SOC column does not match the SOC column on the OCV sheet."
        )

    r0 = np.array([_to_float(r[r0_col_idx]) for r in data_rows])
    return r0, r0_label


def load_cell_model_xlsx(path: str) -> CellModelFromExcel:
    """Loads a SOC-OCV-R0-only cell-model workbook."""
    wb = openpyxl.load_workbook(path, data_only=True)

    missing = [s for s in REQUIRED_SHEETS if s not in wb.sheetnames]
    if missing:
        raise ValueError(
            f"Workbook is missing required sheet(s): {missing}. "
            f"Found sheets: {wb.sheetnames}."
        )

    meta_fields = _read_meta_sheet(wb["Meta"])
    soc, ocv = _read_ocv_sheet(wb["OCV"])
    r0, r0_label = _read_r0_sheet(wb["R0"], expected_soc=soc)

    if "capacity_Ah" not in meta_fields:
        raise ValueError(
            "Meta sheet must define a 'capacity_Ah' row."
        )
    capacity_Ah = float(meta_fields.pop("capacity_Ah"))

    datasheet_raw = {k: meta_fields.pop(k) for k in DATASHEET_FIELDS if k in meta_fields}
    datasheet = apply_default_datasheet(datasheet_raw)

    return CellModelFromExcel(
        soc=soc,
        ocv=ocv,
        r0=r0,
        r0_column_label=r0_label,
        capacity_Ah=capacity_Ah,
        datasheet=datasheet,
        datasheet_meta=meta_fields,
        source_path=path,
    )


def make_sim_cell_model(cm: CellModelFromExcel,
                         eta_fixed: Optional[float] = None) -> dict:
    """Builds a sim-model dict with the same keys/call signatures that
    profile_generation.run_battery_model / tune_battery_model expect,
    so those functions can be reused unchanged.

    Parameters
    ----------
    cm : CellModelFromExcel
        Output of load_cell_model_xlsx.
    eta_fixed : float, optional
        Coulombic efficiency. This schema has no CapacityEta sheet, so
        there's no eta in the workbook. Leave this as None to fall back
        to opts['eta_coulomb_ef'] at simulation time (the same fallback
        _pg_helpers.get_cell_params_fast already uses when a cell model
        has no eta), or pass a fixed value here to bake one in.
    """
    soc_percent = cm.soc * 100.0

    ocv_interp = make_clamped_interp1d(soc_percent, cm.ocv)
    r0_interp = make_clamped_interp1d(soc_percent, cm.r0)

    return {
        "soc_grid_percent": soc_percent,
        "OCV0_interpolant": ocv_interp,
        "OCVrel_interpolant": lambda soc: 0.0,   # no temperature term in this schema
        "Q_interpolant": lambda t: cm.capacity_Ah,
        "R0_interpolant": lambda soc, t: float(r0_interp(soc)),
        "eta_interpolant": (lambda t: eta_fixed) if eta_fixed is not None else None,
    }
